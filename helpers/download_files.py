from openpyxl import Workbook
import io
import base64
from openpyxl.styles import numbers
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from fpdf import FPDF
from datetime import datetime
import plotly.express as px
# Crear un archivo Excel con los datos filtrados
def create_excel_file(df):
    # Crear un libro de trabajo de Excel
    workbook = Workbook()

    # Crear una hoja de cálculo en el libro de trabajo
    sheet = workbook.active

    # Agregar encabezados de columna al archivo Excel
    headers = ['ID', 'ID_sucursal', 'ID_categoria', 'Monto', 'Fecha registrada', 'Descripcion', 'Sucursal', 'Categoria', 'Estado']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    # Agregar datos al archivo Excel
    for row_num, row_data in enumerate(df, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}{row_num}'] = cell_data

    # Establecer el ancho de columna automático
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].auto_size = True

    # Dar formato de peso a la columna 'Monto'
    monto_column = sheet['D']
    for cell in monto_column:
        cell.number_format = '[$$-en-US]#,##0.00'

    # Crear una celda debajo de la columna 'Monto' que haga una autosuma
    autosum_cell = sheet.cell(row=len(df) + 2, column=4)
    autosum_cell.value = f'=SUM(D2:D{len(df) + 1})'
    autosum_cell.number_format = '[$$-en-US]#,##0.00'

    # Crear una tabla en el archivo Excel
    table = Table(displayName="TablaDatos", ref=f'A1:{get_column_letter(len(headers))}{len(df) + 2}')
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    sheet.add_table(table)

    # Guardar el archivo Excel en un objeto de bytes
    excel_data = io.BytesIO()
    workbook.save(excel_data)
    excel_data.seek(0)

    return excel_data

# Descargar el archivo Excel
def download_excel_file(df):
        excel_data = create_excel_file(df)
        b64 = base64.b64encode(excel_data.read()).decode()
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="datos.xlsx"><button style="background-color: #87C489; color: white; padding: 10px 20px; border: none; cursor: pointer; border-radius: 4px; transition: background-color 0.3s ease; text-decoration: none; display: inline-block;">Descargar Excel</button></a>'
        return href

# Función para crear el PDF con mejor diseño y estilos corporativos
def create_pdf(df):
    pdf = FPDF()
    pdf.add_page()

    # Título
    pdf.set_font("Arial", "B", 16)
    pdf.set_text_color(135, 196, 137)  # Color corporativo
    pdf.cell(0, 10, "Reporte de Gastos", ln=True, align='C')
    pdf.ln(10)

    # Fecha y hora de generación del reporte
    pdf.set_font("Arial", "", 10)
    pdf.set_text_color(0, 0, 0)  # Color negro
    now = datetime.now()
    pdf.cell(0, 10, f"Fecha y Hora de Generación: {now.strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
    pdf.ln(10)

    # Resumen estadístico
    pdf.set_font("Arial", "B", 12)
    pdf.set_fill_color(135, 196, 137)  # Color corporativo
    pdf.cell(0, 10, "Resumen Estadístico", ln=True, fill=True)
    pdf.set_font("Arial", "", 10)
    stats = df['monto'].describe()
    for stat, value in stats.items():
        pdf.cell(0, 10, f"{stat.capitalize()}: ${value:.2f}", ln=True)
    pdf.ln(10)

    # Top 5 gastos más altos
    pdf.set_font("Arial", "B", 12)
    pdf.set_fill_color(135, 196, 137)  # Color corporativo
    pdf.cell(0, 10, "Top 5 Gastos más Altos", ln=True, fill=True)
    pdf.set_font("Arial", "", 10)
    top_5 = df.nlargest(5, 'monto')
    for _, row in top_5.iterrows():
        pdf.cell(0, 10, f"Fecha: {row['Fecha']}, Monto: ${row['monto']:.2f}, Categoría: {row['Categoria']}", ln=True)
    pdf.ln(10)

    # Gastos por categoría
    pdf.set_font("Arial", "B", 12)
    pdf.set_fill_color(135, 196, 137)  # Color corporativo
    pdf.cell(0, 10, "Gastos por Categoría", ln=True, fill=True)
    pdf.set_font("Arial", "", 10)
    category_totals = df.groupby('Categoria')['monto'].sum().sort_values(ascending=False)
    for category, total in category_totals.items():
        pdf.cell(0, 10, f"Categoría {category}: ${total:.2f}", ln=True)

    return pdf

# Función para crear el gráfico
def create_plot(df, x, y, title):
    fig = px.bar(df, x=x, y=y, title=title)
    fig.update_layout(xaxis_title='Categoría', yaxis_title='Monto')
    fig.update_traces(marker_color='rgb(158,202,225)', marker_line_color='rgb(8,48,107)', marker_line_width=1.5, opacity=0.6)
    fig.update_layout(title_font=dict(size=20))
    fig.update_layout(showlegend=False)
    fig.update_layout(autosize=False, width=800, height=500)
    return fig
