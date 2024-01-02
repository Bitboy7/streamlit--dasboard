import streamlit as st
import mysql.connector
import os
import time
from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook
import io
import base64
from openpyxl.styles import numbers
import locale
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
load_dotenv()
import plotly.express as px
import plotly.graph_objects as go
st.set_page_config(
    page_title="Datos 📊",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

with open('./style.css') as f:
    st.markdown(f'<style>{f.read()}</style>', unsafe_allow_html=True)

# Configurar variables para la conexión a MySQL
MYSQL_HOST = os.getenv("MYSQL_HOST")
MYSQL_USER = os.getenv("MYSQL_USER")
MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD")
MYSQL_DATABASE = os.getenv("MYSQL_DATABASE")
MYSQL_PORT = os.getenv("MYSQL_PORT")

# Conexión a la base de datos
conn = mysql.connector.connect(
    host=MYSQL_HOST,
    user=MYSQL_USER,
    password=MYSQL_PASSWORD,
    database=MYSQL_DATABASE,
    port=MYSQL_PORT
)
cursor = conn.cursor()
# Check if connection is established
if conn.is_connected():
    st.write("Connected to database")
else:
    st.write("Not connected to database")

# Mostrar datos

st.subheader("Datos 📊")

# Realizar la consulta SQL
with conn:
    cursor.execute("""SELECT r.*, su.nombre AS Sucursal, ca.nombre AS Categoria, es.nombre AS Estado
                      FROM registro r, sucursal su, cat_gastos ca, estado es
                      WHERE r.id_sucursal = su.id
                      AND r.id_cat_gasto = ca.id
                      AND su.id_estado = es.id;""")
    registros = cursor.fetchall()

# Mostrar la tabla en Streamlit
if registros:
    # Crear una lista de diccionarios para los datos de la tabla
    data = [{"ID": registro[0], "id_sucursal": registro[1], "id_cat_gasto": registro[2],
             "monto": registro[3], "fecha registrada": registro[4], "descripcion": registro[5],
             "sucursal": registro[6], "categoria": registro[7], "Estado": registro[8]} for registro in registros]

    # Mostrar la tabla en Streamlit
    expander = st.expander("Filtros.", expanded=False)

    # Obtener los nombres de categoría únicos
    categorias = set([registro["categoria"] for registro in data])
    filtro_categoria = expander.selectbox("Por categoria:", ["Todas"] + list(categorias))

    if filtro_categoria == "Todas":
        filtro_categoria = None

    # Componente de filtrado
    filtro_sucursal = expander.selectbox("Por sucursal:", ["Todas"] + list(
        set([registro["sucursal"] for registro in data])))

    if filtro_sucursal == "Todas":
        filtro_sucursal = None

# Filtrar el dataframe por categoría seleccionada, fecha y sucursal
df_filtrado = pd.DataFrame(data)
if filtro_categoria:
    df_filtrado = df_filtrado[df_filtrado["categoria"] == filtro_categoria]

# Obtener el rango de fechas seleccionado por el usuario
filtro_rango_fecha = expander.selectbox(
    "Por rango de fechas:", ["Día", "Semana", "Mes", "Año"])

if filtro_rango_fecha:
    today = pd.Timestamp.today().normalize()
    if filtro_rango_fecha == "Día":
        start_date = expander.date_input("Fecha inicial:")
        end_date = expander.date_input("Fecha final:", value=today)
    elif filtro_rango_fecha == "Semana":
        start_date = expander.date_input(
            "Semana inicial:", value=today - pd.DateOffset(weeks=1))
        end_date = expander.date_input("Semana final:", value=today)
    elif filtro_rango_fecha == "Mes":
        start_date = expander.date_input(
            "Mes inicial:", value=today - pd.DateOffset(months=1))
        end_date = expander.date_input("Mes final:", value=today)
    elif filtro_rango_fecha == "Año":
        start_date = expander.date_input(
            "Año inicial:", value=today - pd.DateOffset(years=1))
        end_date = expander.date_input("Año final:", value=today)

    # Filtrar el dataframe por rango de fechas seleccionado
    if start_date and end_date:
        df_filtrado = df_filtrado[
            (pd.to_datetime(df_filtrado["fecha registrada"]).dt.date >= start_date) &
            (pd.to_datetime(
                df_filtrado["fecha registrada"]).dt.date <= end_date)
        ]
if filtro_sucursal:
    df_filtrado = df_filtrado[df_filtrado["sucursal"] == filtro_sucursal]

# Formatear la columna "monto" como moneda
locale.setlocale(locale.LC_ALL, '')  # Set the locale to the user's default

df_filtrado["monto"] = df_filtrado["monto"].map(
    lambda x: locale.currency(x, grouping=True))

# Mostrar el dataframe filtrado
expander.dataframe(df_filtrado)

# Obtener el tipo de estadística seleccionada por el usuario
filtro_estadistica = expander.selectbox(
    "Tipo de estadística:", ["Suma", "Promedio", "Mínimo", "Máximo"])

if filtro_estadistica:
    # Convertir la columna "monto" a números
    df_filtrado["monto"] = pd.to_numeric(
        df_filtrado["monto"].str.replace('[^\d.]', ''), errors='coerce')

    if filtro_estadistica == "Suma":
        # Calcular la suma de la columna "monto"
        suma_monto = df_filtrado["monto"].sum()
        expander.write(f"Suma de la columna 'monto': ${suma_monto} pesos.")
    elif filtro_estadistica == "Promedio":
        # Calcular el promedio de la columna "monto"
        promedio_monto = df_filtrado["monto"].mean()
        expander.write(f"Promedio de la columna 'monto': ${promedio_monto} pesos.")
    elif filtro_estadistica == "Mínimo":
        # Obtener el valor mínimo de la columna "monto"
        minimo_monto = df_filtrado["monto"].min()
        expander.markdown(
            f"Valor mínimo de la columna 'monto': {minimo_monto}")
    elif filtro_estadistica == "Máximo":
        # Obtener el valor máximo de la columna "monto"
        maximo_monto = df_filtrado["monto"].max()
        expander.markdown(
            f"Valor máximo de la columna 'monto': {maximo_monto}")

# Generar el archivo Excel a partir del dataframe filtrado
@st.cache_data      
def generate_excel(df_filtrado):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(['ID', 'ID_sucursal', 'ID_categoria', 'Monto',
                     'Fecha registrada', 'Descripcion', 'Sucursal', 'Categoria', 'Estado'])
    for _, row in df_filtrado.iterrows():
        worksheet.append(list(row))

    # Add totals row with sum formula in 'monto' column
    last_row = worksheet.max_row
    totals_row = [f"=SUM(D2:D{last_row})"] + [""] * (worksheet.max_column - 1)
    worksheet.append(totals_row)

    # Apply formatting to totals row
    totals_row_number = last_row + 1
    for column in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=totals_row_number, column=column)
        cell.alignment = Alignment(horizontal='right')

    # Create table and apply style to the table
    table_range = f"A1:{get_column_letter(worksheet.max_column)}{totals_row_number}"
    table = Table(displayName="Table1", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    worksheet.add_table(table)
    # Format column 'monto' as currency
    currency_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    worksheet.column_dimensions['D'].number_format = currency_format
    return workbook

# Devuelve el contenido del archivo Excel como una cadena binaria

@st.cache_data
def get_binary_content(workbook):
    binary_content = io.BytesIO()
    workbook.save(binary_content)
    return binary_content.getvalue()


if expander.button("Generar Excel", key="generar_excel", type="primary"):
    # Generar el archivo Excel a partir del dataframe filtrado
    workbook = generate_excel(df_filtrado)

    # Obtener el contenido binario del archivo Excel
    excel_content = get_binary_content(workbook)

    # Codificar el contenido binario en base64
    excel_base64 = base64.b64encode(excel_content).decode('utf-8')

    # Generar el enlace de descarga del archivo Excel
    href = f'<a href="data:application/octet-stream;base64,{excel_base64}" download="datos.xlsx" style="color: white; text-decoration: none; font-weight: bold; background-color: green; padding: 0.5rem 1rem; text-decoration: none;transition: all 0.3s ease-out;border-radius: 1rem; margin: 8px;">Descargar Excel</a>'
    expander.markdown(href, unsafe_allow_html=True)
    progress_text = "Operation in progress..."
    my_bar = st.progress(0, text=progress_text)

    for percent_complete in range(100):
      time.sleep(0.01)
    my_bar.progress(percent_complete + 1, text=progress_text)
    time.sleep(1)
    my_bar.empty()
    time.sleep(2.5)
    st.toast(f"Archivo Excel generado con éxito")
    
# Crear gráfico interactivo con Plotly
fig = px.bar(df_filtrado, x='sucursal', y='monto', title="Gráfico por Categoría.",
             hover_data=['categoria', 'monto'], color='categoria', text_auto=True ,
             labels={'costo':'categoria'}, height=500)

# Crear gráfico interactivo con Plotly
fig2 = px.bar(df_filtrado, x='categoria', y='monto',title="Gráfico por Sucursal.",
             hover_data=['sucursal', 'monto'], color='sucursal', text_auto=True,
             labels={'gastos':'Gráfico de Montos por Sucursal.'}, height=500)


fig3 = px.pie(df_filtrado, values='monto', names='categoria', title='Gastos por Categoría')


fig4 = px.pie(df_filtrado, values='monto', names='sucursal', title='Gastos por Sucursal')


st.subheader("Gráficas.")
# Obtener el tipo de gráfico seleccionado por el usuario
filtro_grafico = st.selectbox(
    "Tipo de gráfico:", ["Gráfico de Montos por Categoría (Bar)", "Gráfico de Montos por Sucursal (Bar)", "Gráfico de Montos por Categoría (Pie)", "Gráfico de Montos por Sucursal (Pie)"])

if filtro_grafico == "Gráfico de Montos por Categoría (Bar)":
    # Mostrar gráfico de Montos por Categoría en Streamlit
    st.plotly_chart(fig)
elif filtro_grafico == "Gráfico de Montos por Sucursal (Bar)":
    # Mostrar gráfico de Montos por Sucursal en Streamlit
    st.plotly_chart(fig2)
elif filtro_grafico == "Gráfico de Montos por Categoría (Pie)":
    # Mostrar gráfico de Montos por Categoría (Pie) en Streamlit
    st.plotly_chart(fig3)
elif filtro_grafico == "Gráfico de Montos por Sucursal (Pie)":
    st.plotly_chart(fig4)    
